from flask import Flask, request, render_template, send_file, jsonify, Response
import qrcode
import barcode
from barcode.writer import ImageWriter
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
import io
import os
import base64

app = Flask(__name__, template_folder='templates')

# Functions to generate QR code and barcode as data URI
# (same as previous versions)
def generate_qr_code(data):
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=5,
        border=4,
    )
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill='black', back_color='white')
    img_bytes = io.BytesIO()
    img.save(img_bytes, format='PNG')
    img_bytes.seek(0)
    b64_image = base64.b64encode(img_bytes.read()).decode('utf-8')
    return b64_image

def generate_barcode(data):
    EAN = barcode.get_barcode_class('code128')
    ean = EAN(data, writer=ImageWriter())
    options = {
        'module_width': 0.15,  # Adjust the width of each module (narrower bars)
        'module_height': 6.0,  # Adjust the height of the barcode
        'quiet_zone': 1.0,  # Adjust the quiet zone if needed
        'text_distance': 5.0,  # Distance between the barcode and the human-readable text
        'font_size': 5.0,  # Reduce font size to make the text smaller
    }
    barcode_bytes = io.BytesIO()
    ean.write(barcode_bytes,options=options)
    barcode_bytes.seek(0)
    b64_image = base64.b64encode(barcode_bytes.read()).decode('utf-8')
    return b64_image

from PIL import Image as PILImage

def generate_qr_code_bytesio(data):
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill='black', back_color='white')
    img_bytes = io.BytesIO()
    img.save(img_bytes, format='PNG')
    img_bytes.seek(0)
    return img_bytes

def generate_barcode_bytesio(data):
    EAN = barcode.get_barcode_class('code128')
    ean = EAN(data, writer=ImageWriter())
    barcode_bytes = io.BytesIO()
    ean.write(barcode_bytes)
    barcode_bytes.seek(0)
    return barcode_bytes


def resize_image(img_bytes, max_size):
    """Resize PIL image to fit within Excel cell."""
    img = PILImage.open(img_bytes)
    # Calculate the scaling factor, keeping aspect ratio
    factor = min(max_size[0] / img.width, max_size[1] / img.height)
    # Resize image using LANCZOS, which is recommended for high-quality downscaling
    resized_img = img.resize((int(img.width * factor), int(img.height * factor)), PILImage.LANCZOS)
    # Save resized image to BytesIO
    new_img_bytes = io.BytesIO()
    resized_img.save(new_img_bytes, format='PNG')
    new_img_bytes.seek(0)
    return new_img_bytes

def create_excel_with_codes(data_list):
    wb = Workbook()
    ws = wb.active

    # Set column widths for better visibility
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 40
    
    # Insert headers
    headers = ["Name", "Phone No", "Email ID", "Bar Code", "QR Code"]
    ws.append(headers)
    
    # Insert data and QR/Barcode images
    for data in data_list:
        name = data['name']
        phone = data['phone']
        email = data['email']

        # Generate images
        qr_img_bytes = generate_qr_code_bytesio(email)
        barcode_img_bytes = generate_barcode_bytesio(phone)

        # Load images into openpyxl
        qr_img = ExcelImage(qr_img_bytes)
        barcode_img = ExcelImage(barcode_img_bytes)

        # Append data to row
        ws.append([name, phone, email])

        # Insert images
        cell_coord_qr = f'E{ws._current_row}'
        cell_coord_bar = f'D{ws._current_row}'
        ws.add_image(barcode_img, cell_coord_bar)
        ws.add_image(qr_img, cell_coord_qr)

        # Adjust row height to fit images
        ws.row_dimensions[ws._current_row].height = 250  # Adjust based on your image size

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST' and request.headers.get('Content-Type') == 'application/json':
        data = request.get_json()
        # Process JSON data here if needed, like generating the Excel file dynamically based on data
        # Placeholder response for AJAX success
        return jsonify({'success': True})
    
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit_data():
    name = request.form['name']
    phone = request.form['phone']
    email = request.form['email']

    if not (name and phone and email):
        return jsonify({'error': 'Missing data'}), 400

    qr_code = generate_qr_code(email)
    barcode = generate_barcode(phone)

    return jsonify(name=name, phone=phone, email=email, qrCode=f"data:image/png;base64,{qr_code}", barcode=f"data:image/png;base64,{barcode}")

@app.route('/download_excel', methods=['POST'])
def download_excel():
    # Assuming data is sent as JSON in the request
    data = request.get_json().get('data')

    if not data:
        return "No data provided", 400

    output = create_excel_with_codes(data)  # Assuming create_excel_with_codes returns a BytesIO object of the Excel file

    return send_file(
        output,
        as_attachment=True,
        download_name="output.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == '__main__':
    app.run(debug=True)
